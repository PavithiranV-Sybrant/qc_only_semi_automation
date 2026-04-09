import io
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

_PURPLE_FILL = PatternFill(start_color="B19CD9", end_color="B19CD9", fill_type="solid")


@st.cache_data
def load_file(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(uploaded_file, dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(uploaded_file, dtype=str, keep_default_na=False)
    return df.replace("", float("nan"))


def render_preview(df: pd.DataFrame):
    # ── Top-level metrics ────────────────────────────────────────────
    total_cells = len(df) * len(df.columns)
    null_cells  = int(df.isnull().sum().sum())
    dup_rows    = int(df.duplicated().sum())
    fill_pct    = round((1 - null_cells / total_cells) * 100, 1) if total_cells else 0.0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Rows",            f"{len(df):,}")
    c2.metric("Columns",         len(df.columns))
    c3.metric("Total Nulls",     f"{null_cells:,}")
    c4.metric("Duplicate Rows",  f"{dup_rows:,}")
    c5.metric("Fill Rate",       f"{fill_pct}%")

    st.divider()

    tab_table, tab_quality, tab_explorer = st.tabs(
        ["Table View", "Data Quality", "Column Explorer"]
    )

    # ── Tab 1: raw table ────────────────────────────────────────────
    with tab_table:
        st.dataframe(df.head(200), use_container_width=True, height=400)

    # ── Tab 2: data quality ─────────────────────────────────────────
    with tab_quality:
        null_counts   = df.isnull().sum()
        unique_counts = df.nunique()
        non_null      = df.count()

        quality_df = pd.DataFrame({
            "Column":    df.columns,
            "Non-Null":  non_null.values,
            "Null":      null_counts.values,
            "Null %":    (null_counts / len(df) * 100).round(1).values,
            "Unique":    unique_counts.values,
            "Unique %":  (unique_counts / len(df) * 100).round(1).values,
            "Sample":    [df[c].dropna().iloc[0] if df[c].notna().any() else "" for c in df.columns],
        })

        def _color_null_pct(series):
            def _cell(val):
                if val >= 50: return "background-color: #f4999a"
                if val >= 20: return "background-color: #fad4a6"
                if val > 0:   return "background-color: #fef3cd"
                return ""
            return [_cell(v) for v in series]

        st.subheader("Column Quality Summary")
        st.dataframe(
            quality_df.style.apply(_color_null_pct, subset=["Null %"]),
            use_container_width=True,
            hide_index=True,
        )


    # ── Tab 3: column explorer ──────────────────────────────────────
    with tab_explorer:
        col = st.selectbox("Choose a column to explore", df.columns, key="explorer_col")
        if col is None:
            st.info("Select a column above.")
            return

        series    = df[col]
        non_null  = series.dropna()
        null_cnt  = int(series.isnull().sum())
        unique_n  = int(series.nunique())

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Non-Null",  f"{len(non_null):,}")
        m2.metric("Null",      f"{null_cnt:,}")
        m3.metric("Unique",    f"{unique_n:,}")
        m4.metric("Null %",    f"{round(null_cnt / len(df) * 100, 1)}%")

        st.divider()

        # Value distribution
        top_n = st.selectbox(
            "Show top N values",
            options=[5, 10, 20, 50, 100],
            index=1,
            key="explorer_topn",
        )
        vc     = series.value_counts().head(top_n)
        vc_df  = vc.reset_index()
        vc_df.columns = ["Value", "Count"]
        vc_df["% of Rows"] = (vc_df["Count"] / len(df) * 100).round(1)

        left, right = st.columns([2, 3])
        with left:
            st.markdown(f"**Top {top_n} Values**")
            st.dataframe(vc_df, use_container_width=True, hide_index=True, height=360)
        with right:
            st.markdown(f"**Distribution (top {top_n})**")
            st.bar_chart(vc, height=340)

        st.divider()

        # Interactive filter
        st.markdown("**Filter rows by value**")
        all_vals = sorted(series.dropna().astype(str).unique().tolist())
        display_vals = all_vals[:300]  # cap options to avoid UI overload

        selected = st.multiselect(
            f"Select values in  '{col}'",
            options=display_vals,
            key="explorer_filter",
        )

        if selected:
            mask     = series.astype(str).isin(selected)
            filtered = df[mask]
            st.markdown(f"**{len(filtered):,} of {len(df):,} rows** match")
            st.dataframe(filtered, use_container_width=True, height=300)
        else:
            st.caption("No filter applied — select one or more values above to narrow down rows.")


def _write_excel_with_purple_headers(df: pd.DataFrame, original_cols: set, buf: io.BytesIO):
    """Write df to buf as .xlsx, coloring header cells purple for newly added columns."""
    df.to_excel(buf, index=False)
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active
    for cell in ws[1]:
        if cell.value and cell.value not in original_cols:
            cell.fill = _PURPLE_FILL
    buf.seek(0)
    buf.truncate()
    wb.save(buf)


def render_download(df: pd.DataFrame, original_cols: set = None):
    buf = io.BytesIO()
    if original_cols:
        _write_excel_with_purple_headers(df, original_cols, buf)
    else:
        df.to_excel(buf, index=False)
    st.download_button(
        label="Download QC Output",
        data=buf.getvalue(),
        file_name="qc_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
