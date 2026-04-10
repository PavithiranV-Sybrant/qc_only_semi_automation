import io
import uuid
import zipfile
import openpyxl
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import APIRouter, File, UploadFile, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel

from backend.session_store import create_session, update_session, get_session
from backend.pipeline_executor import execute_pipeline
from backend.file_store import save_file as _persist_file, get_file as _get_stored_file
from backend import job_queue

router = APIRouter()
_PURPLE = openpyxl.styles.PatternFill(fill_type="solid", fgColor="B19CD9")

# batch_id → { files: [...], status, cancelled, started_at, name }
_batch_store: dict[str, dict] = {}


# ─── helpers ────────────────────────────────────────────────────────────────

def _clean(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        mask = df[col].str.strip().eq("") if df[col].dtype == object else pd.Series(False, index=df.index)
        df.loc[mask, col] = pd.NA
    return df


def _generate_excel(df: pd.DataFrame, original_cols: list) -> bytes:
    new_cols = set(df.columns) - set(original_cols)
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.append(list(df.columns))
    for cell in ws[1]:
        if cell.value in new_cols:
            cell.fill = _PURPLE
    for row in df.itertuples(index=False, name=None):
        ws.append([None if (isinstance(v, float) and v != v) else v for v in row])
    wb.save(buf)
    return buf.getvalue()


# ─── public store accessors (used by background router) ──────────────────────

def list_batches() -> list:
    """Return all batches sorted newest-first."""
    batches = [{"id": bid, **b} for bid, b in _batch_store.items()]
    return sorted(batches, key=lambda b: b.get("started_at") or "", reverse=True)


def dismiss_batch(bid: str) -> bool:
    if bid in _batch_store:
        del _batch_store[bid]
        return True
    return False


# ─── models ─────────────────────────────────────────────────────────────────

class BatchRunRequest(BaseModel):
    batch_id:       str
    column_mapping: dict[str, Any]
    step_toggles:   dict[str, bool]
    thresholds:     dict[str, float]


# ─── background task ────────────────────────────────────────────────────────

def _run_file(batch_id: str, session_id: str, req: BatchRunRequest):
    batch      = _batch_store.get(batch_id, {})
    file_entry = next((f for f in batch.get("files", []) if f["session_id"] == session_id), None)
    sess       = get_session(session_id)

    if not sess:
        if file_entry:
            file_entry["status"] = "error"
            file_entry["error"]  = "Session not found."
        _check_batch_done(batch_id)
        return

    if file_entry:
        file_entry["status"] = "running"

    def progress_cb(step_index, total_steps, label):
        if file_entry:
            file_entry["step_index"]   = step_index
            file_entry["total_steps"]  = total_steps
            file_entry["current_step"] = label

    def cancel_check():
        return batch.get("cancelled", False)

    try:
        df = sess["df_original"].copy()
        df_out, _, elapsed = execute_pipeline(
            df, req.column_mapping, req.step_toggles, req.thresholds,
            progress_cb=progress_cb,
            cancel_check=cancel_check,
        )

        if cancel_check():
            if file_entry:
                file_entry["status"] = "cancelled"
        else:
            original_cols = sess["original_columns"]
            excel_bytes   = _generate_excel(df_out, original_cols)
            update_session(session_id, {"excel_bytes": excel_bytes, "df_working": df_out})

            # Persist to disk; store id for later download
            storage_file_id = None
            try:
                rec = _persist_file(sess.get("file_name", "output.xlsx"), excel_bytes)
                storage_file_id = rec["id"]
            except Exception:
                pass

            if file_entry:
                file_entry["status"]          = "done"
                file_entry["elapsed"]         = round(elapsed, 1)
                file_entry["download_ready"]  = True
                file_entry["storage_file_id"] = storage_file_id

    except Exception as e:
        if file_entry:
            file_entry["status"] = "error"
            file_entry["error"]  = str(e)

    _check_batch_done(batch_id)


def _check_batch_done(batch_id: str):
    batch = _batch_store.get(batch_id, {})
    if all(f["status"] in ("done", "error", "cancelled") for f in batch.get("files", [])):
        batch["status"] = "cancelled" if batch.get("cancelled") else "done"


# ─── routes ─────────────────────────────────────────────────────────────────

@router.post("/batch/upload")
async def batch_upload(files: list[UploadFile] = File(...)):
    batch_id      = str(uuid.uuid4())
    file_sessions = []

    for file in files:
        name = file.filename or ""
        ext  = name.rsplit(".", 1)[-1].lower()
        if ext not in ("xlsx", "csv"):
            continue

        content = await file.read()
        try:
            if ext == "xlsx":
                df = pd.read_excel(io.BytesIO(content), dtype=str)
            else:
                df = pd.read_csv(io.BytesIO(content), dtype=str)
            df = _clean(df)
        except Exception:
            continue

        sid = create_session()
        update_session(sid, {
            "df_original":      df,
            "original_columns": list(df.columns),
            "file_name":        name,
        })
        file_sessions.append({
            "session_id":     sid,
            "file_name":      name,
            "rows":           len(df),
            "columns":        len(df.columns),
            "column_names":   list(df.columns),
            "status":         "ready",
            "step_index":     0,
            "total_steps":    0,
            "current_step":   "",
            "download_ready": False,
            "storage_file_id": None,
        })

    if not file_sessions:
        raise HTTPException(400, "No valid .xlsx or .csv files found in upload.")

    _batch_store[batch_id] = {
        "files":      file_sessions,
        "status":     "ready",
        "cancelled":  False,
        "started_at": None,
        "name":       None,
    }
    return {"batch_id": batch_id, "file_count": len(file_sessions), "files": file_sessions}


@router.post("/batch/run")
def run_batch(req: BatchRunRequest):
    batch = _batch_store.get(req.batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found.")

    batch["status"]     = "running"
    batch["cancelled"]  = False
    batch["started_at"] = datetime.now(timezone.utc).isoformat()

    # Auto-generate display name
    names = [f["file_name"] for f in batch["files"]]
    if len(names) == 1:
        batch["name"] = names[0]
    else:
        batch["name"] = f"{names[0]} + {len(names) - 1} more"

    for f in batch["files"]:
        f["status"]          = "pending"
        f["step_index"]      = 0
        f["total_steps"]     = 0
        f["current_step"]    = ""
        f["download_ready"]  = False
        f["storage_file_id"] = None

    def task():
        job_queue.mark_running(req.batch_id)
        for f in batch["files"]:
            if batch.get("cancelled"):
                if f["status"] == "pending":
                    f["status"] = "cancelled"
                continue
            _run_file(req.batch_id, f["session_id"], req)  # blocking — serial execution
        _check_batch_done(req.batch_id)
        if batch.get("cancelled"):
            job_queue.mark_cancelled(req.batch_id)
        else:
            job_queue.mark_done(req.batch_id)

    job_queue.enqueue(task, {"id": req.batch_id, "type": "batch", "label": batch["name"]})
    return {"status": "queued", "file_count": len(batch["files"])}


@router.get("/batch/status/{batch_id}")
def batch_status(batch_id: str):
    batch = _batch_store.get(batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found.")
    return batch


@router.post("/batch/cancel/{batch_id}")
def cancel_batch(batch_id: str):
    batch = _batch_store.get(batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found.")
    batch["cancelled"] = True
    batch["status"]    = "cancelled"
    for f in batch["files"]:
        if f["status"] in ("pending", "running"):
            f["status"] = "cancelled"
    return {"status": "cancelling"}


@router.get("/batch/download/{session_id}")
def batch_download_file(session_id: str):
    sess = get_session(session_id)
    if not sess or "excel_bytes" not in sess:
        raise HTTPException(404, "File not ready for download.")
    file_name = sess.get("file_name", "output.xlsx")
    stem = file_name.rsplit(".", 1)[0]
    return Response(
        content=sess["excel_bytes"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="qc_{stem}.xlsx"'},
    )


@router.get("/batch/download-all/{batch_id}")
def batch_download_all(batch_id: str):
    """Stream a ZIP archive of all completed files in the batch."""
    batch = _batch_store.get(batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found.")

    done_files = [f for f in batch["files"] if f["status"] == "done"]
    if not done_files:
        raise HTTPException(404, "No completed files in this batch yet.")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in done_files:
            excel_bytes = None
            # Prefer in-memory session (fast path)
            sess = get_session(f["session_id"])
            if sess and "excel_bytes" in sess:
                excel_bytes = sess["excel_bytes"]
            # Fallback: read from persistent disk storage
            elif f.get("storage_file_id"):
                rec = _get_stored_file(f["storage_file_id"])
                if rec:
                    excel_bytes = Path(rec["path"]).read_bytes()

            if excel_bytes:
                stem = f["file_name"].rsplit(".", 1)[0]
                zf.writestr(f"qc_{stem}.xlsx", excel_bytes)

    return Response(
        content=buf.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="batch_{batch_id[:8]}_output.zip"'},
    )
