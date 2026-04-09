import io

import openpyxl
from openpyxl.styles import PatternFill
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from backend.session_store import get_session

router = APIRouter()

_PURPLE = PatternFill(fill_type="solid", fgColor="B19CD9")


@router.get("/download/{session_id}")
def download(session_id: str):
    sess = get_session(session_id)
    if not sess or "df_working" not in sess:
        raise HTTPException(404, "Session not found or pipeline not run yet.")

    df              = sess["df_working"]
    original_cols   = set(sess.get("original_columns", []))
    new_cols        = set(df.columns) - original_cols

    buf = io.BytesIO()
    with openpyxl.Workbook() as wb:
        ws = wb.active
        ws.append(list(df.columns))
        for cell in ws[1]:
            if cell.value in new_cols:
                cell.fill = _PURPLE
        for row in df.itertuples(index=False, name=None):
            ws.append(list(row))
        wb.save(buf)

    buf.seek(0)
    fname = sess.get("file_name", "output").rsplit(".", 1)[0]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}_qc_output.xlsx"'},
    )
