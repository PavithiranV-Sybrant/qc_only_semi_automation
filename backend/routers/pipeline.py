import io
import openpyxl
from openpyxl.styles import PatternFill
from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Any

from backend.session_store import get_session, update_session
from backend.job_store import create_job, get_job, update_job
from backend.pipeline_executor import execute_pipeline
from backend.file_store import save_file as _persist_file
from backend import job_queue

router  = APIRouter()
_PURPLE = PatternFill(fill_type="solid", fgColor="B19CD9")


class RunRequest(BaseModel):
    session_id:     str
    column_mapping: dict[str, Any]
    step_toggles:   dict[str, bool]
    thresholds:     dict[str, float]


def _generate_excel(df, original_cols):
    """Build Excel bytes with purple headers for new columns."""
    new_cols = set(df.columns) - set(original_cols)
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.append(list(df.columns))
    for cell in ws[1]:
        if cell.value in new_cols:
            cell.fill = _PURPLE
    for row in df.itertuples(index=False, name=None):
        ws.append([None if (isinstance(v, float) and v != v) else v for v in row])
    wb.save(buf)
    return buf.getvalue()


def _run_pipeline_task(job_id: str, req: RunRequest):
    """Runs in a background thread. Updates job store with progress and results."""
    sess = get_session(req.session_id)
    if not sess:
        update_job(job_id, {"status": "error", "error": "Session not found. Please re-upload the file."})
        return

    def progress_cb(step_index, total_steps, label):
        update_job(job_id, {
            "status":       "running",
            "step_index":   step_index,
            "total_steps":  total_steps,
            "current_step": label,
        })

    try:
        def cancel_check():
            return get_job(job_id).get("cancelled", False)

        df = sess["df_original"].copy()
        df_out, results, elapsed = execute_pipeline(
            df, req.column_mapping, req.step_toggles, req.thresholds,
            progress_cb=progress_cb,
            cancel_check=cancel_check,
        )

        if cancel_check():
            update_job(job_id, {"status": "cancelled", "current_step": "Cancelled by user."})
            return

        original_cols = sess["original_columns"]
        new_cols      = [c for c in df_out.columns if c not in set(original_cols)]

        new_cols_summary = []
        distributions    = {}
        for col in new_cols:
            total     = len(df_out)
            populated = int(df_out[col].notna().sum())
            vc        = df_out[col].value_counts(dropna=True)
            top_val   = str(vc.index[0]) if len(vc) else ""
            new_cols_summary.append({
                "column":    col,
                "populated": populated,
                "null":      total - populated,
                "unique":    int(df_out[col].nunique(dropna=True)),
                "top_value": top_val,
            })
            distributions[col] = {str(k): int(v) for k, v in vc.head(20).items()}

        update_session(req.session_id, {"df_working": df_out})

        # Pre-generate Excel so download is instant
        update_job(job_id, {"status": "preparing_download", "current_step": "Preparing download file..."})
        excel_bytes = _generate_excel(df_out, original_cols)
        update_session(req.session_id, {"excel_bytes": excel_bytes})

        # Persist to disk for later re-download; capture storage id for background page
        storage_file_id = None
        try:
            rec = _persist_file(sess.get("file_name", "output.xlsx"), excel_bytes)
            storage_file_id = rec["id"]
        except Exception:
            pass  # storage failure must not abort the pipeline

        update_job(job_id, {
            "status":           "done",
            "results":          results,
            "elapsed_total":    elapsed,
            "new_columns":      new_cols,
            "new_cols_summary": new_cols_summary,
            "distributions":    distributions,
            "rows":             len(df_out),
            "download_ready":   True,
            "storage_file_id":  storage_file_id,
        })
    except Exception as e:
        update_job(job_id, {"status": "error", "error": str(e)})


@router.post("/run-pipeline")
def run_pipeline(req: RunRequest):
    """Enqueue pipeline job. Returns job_id immediately; runs when worker is free."""
    sess = get_session(req.session_id)
    if not sess:
        raise HTTPException(404, "Session not found. Please re-upload the file.")
    job_id = create_job(req.session_id, file_name=sess.get("file_name", ""))

    def task():
        job_queue.mark_running(job_id)
        try:
            _run_pipeline_task(job_id, req)
        finally:
            final = get_job(job_id)
            if final:
                s = final.get("status", "")
                if s == "error":
                    job_queue.mark_error(job_id, final.get("error", ""))
                elif s == "cancelled":
                    job_queue.mark_cancelled(job_id)
                else:
                    job_queue.mark_done(job_id)

    job_queue.enqueue(task, {"id": job_id, "type": "single",
                              "label": sess.get("file_name", "file")})
    return {"job_id": job_id, "status": "queued"}


@router.get("/pipeline-status/{job_id}")
def pipeline_status(job_id: str):
    """Poll this endpoint to get live progress and final results."""
    job = get_job(job_id)
    if not job:
        raise HTTPException(404, "Job not found.")
    return job


@router.post("/cancel-pipeline/{job_id}")
def cancel_pipeline(job_id: str):
    """Signal the running pipeline to stop after the current step."""
    job = get_job(job_id)
    if not job:
        raise HTTPException(404, "Job not found.")
    update_job(job_id, {"cancelled": True, "status": "cancelled", "current_step": "Cancelling..."})
    return {"status": "cancelling"}
