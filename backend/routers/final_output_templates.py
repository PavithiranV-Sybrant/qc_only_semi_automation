import io
import json
import math
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
import openpyxl.styles
import pandas as pd
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from fuzzywuzzy import fuzz
from pydantic import BaseModel

router = APIRouter()

_FINAL_TEMPLATES_DIR = Path(__file__).parent.parent.parent / "instructions" / "final_templates"
_MAX_FILE_BYTES = 500 * 1024 * 1024


# ── Pydantic models ──────────────────────────────────────────────────────────

class ColumnEntry(BaseModel):
    name: str
    argb: str


class FinalTemplatePayload(BaseModel):
    comment: Optional[str] = ""
    columns: list[ColumnEntry]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _normalize(name: str) -> str:
    """Normalize a column name for fuzzy comparison: lowercase, collapse separators → underscore."""
    s = name.strip().lower()
    s = re.sub(r"[\s\-\./\\]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def _extract_argb(fg_color) -> str:
    """Extract ARGB string from an openpyxl fgColor object."""
    try:
        if fg_color.type == "rgb":
            return fg_color.rgb          # e.g. "FF00B0F0"
        elif fg_color.type == "theme":
            return f"theme:{fg_color.theme}"
        elif fg_color.type == "indexed":
            return f"indexed:{fg_color.indexed}"
    except Exception:
        pass
    return "00000000"


def _load_df(content: bytes, filename: str) -> pd.DataFrame:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "xlsx":
        return pd.read_excel(io.BytesIO(content), dtype=str)
    return pd.read_csv(io.BytesIO(content), dtype=str)


def _load_df_headers(content: bytes, filename: str) -> list[str]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "xlsx":
        df = pd.read_excel(io.BytesIO(content), nrows=0, dtype=str)
    else:
        df = pd.read_csv(io.BytesIO(content), nrows=0, dtype=str)
    return list(df.columns)


# ── IMPORTANT: static paths must be registered before {name} param routes ───

@router.post("/final-templates/extract-headers")
async def extract_headers(file: UploadFile = File(...)):
    """Upload a golden .xlsx file and return its column names with header background colors."""
    fname = file.filename or ""
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""
    if ext != "xlsx":
        raise HTTPException(400, "Only .xlsx files carry header colors. Please upload an Excel (.xlsx) file.")

    content = await file.read()
    if len(content) > _MAX_FILE_BYTES:
        raise HTTPException(413, "File too large (max 500 MB).")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if header_row is None:
            raise HTTPException(400, "Excel file appears to be empty.")
        columns = []
        for cell in header_row:
            val = cell.value
            if val is None or str(val).strip() == "":
                continue
            argb = _extract_argb(cell.fill.fgColor)
            columns.append({"name": str(val), "argb": argb})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Could not read Excel headers: {e}")

    return {"columns": columns}


@router.post("/final-templates/check")
async def check_file(
    file: UploadFile = File(...),
    template_name: str = Form(...),
):
    """Compare uploaded file columns against a saved final template and return a mapping analysis."""
    tpl_path = _FINAL_TEMPLATES_DIR / f"{template_name}.json"
    if not tpl_path.exists():
        raise HTTPException(404, f"Template '{template_name}' not found.")

    tpl_data = json.loads(tpl_path.read_text(encoding="utf-8"))
    tpl_cols = [c["name"] for c in tpl_data.get("columns", [])]

    content = await file.read()
    fname = file.filename or "upload"

    try:
        file_cols = _load_df_headers(content, fname)
    except Exception as e:
        raise HTTPException(400, f"Could not read file headers: {e}")

    file_col_set = set(file_cols)
    matched_set: set[str] = set()
    mappings = []

    for tpl_col in tpl_cols:
        # 1. Exact match
        if tpl_col in file_col_set and tpl_col not in matched_set:
            mappings.append({
                "template_col": tpl_col,
                "file_col":     tpl_col,
                "action":       "exact",
                "confidence":   100,
                "suggestions":  [],
            })
            matched_set.add(tpl_col)
            continue

        # 2. Normalized match (e.g. "first name" → "first_name")
        norm_tpl = _normalize(tpl_col)
        norm_match = None
        for fc in file_cols:
            if fc not in matched_set and _normalize(fc) == norm_tpl:
                norm_match = fc
                break

        if norm_match:
            mappings.append({
                "template_col": tpl_col,
                "file_col":     norm_match,
                "action":       "auto_rename",
                "confidence":   100,
                "suggestions":  [norm_match],
            })
            matched_set.add(norm_match)
            continue

        # 3. Fuzzy match
        top3: list[str] = []
        unmatched = [fc for fc in file_cols if fc not in matched_set]
        if unmatched:
            scored = []
            for fc in unmatched:
                score = max(
                    fuzz.ratio(tpl_col.lower(), fc.lower()),
                    fuzz.token_sort_ratio(tpl_col.lower(), fc.lower()),
                    fuzz.ratio(norm_tpl, _normalize(fc)),
                )
                scored.append((fc, score))
            scored.sort(key=lambda x: x[1], reverse=True)
            top3 = [s[0] for s in scored[:3]]

            if scored[0][1] >= 70:
                best_fc, best_score = scored[0]
                mappings.append({
                    "template_col": tpl_col,
                    "file_col":     best_fc,
                    "action":       "fuzzy_suggest",
                    "confidence":   best_score,
                    "suggestions":  top3,
                })
                matched_set.add(best_fc)
                continue

        mappings.append({
            "template_col": tpl_col,
            "file_col":     None,
            "action":       "missing",
            "confidence":   0,
            "suggestions":  top3,
        })

    # Post-process: if all non-missing are exact/auto_rename, check if order also needs fixing
    non_missing = [m for m in mappings if m["action"] != "missing"]
    if non_missing and all(m["action"] in ("exact", "auto_rename") for m in non_missing):
        file_col_idx = {fc: i for i, fc in enumerate(file_cols)}
        matched_positions = [file_col_idx.get(m["file_col"], 999) for m in non_missing]
        if matched_positions != sorted(matched_positions):
            for m in non_missing:
                if m["action"] == "exact":
                    m["action"] = "reorder"

    # Determine overall status
    actions = {m["action"] for m in mappings}
    if not actions or actions <= {"exact"}:
        overall = "exact_match"
    elif actions <= {"exact", "reorder"}:
        overall = "reorder_only"
    elif "missing" in actions:
        overall = "missing_columns"
    else:
        overall = "needs_work"

    extra_cols = [fc for fc in file_cols if fc not in matched_set]

    return {
        "overall_status":  overall,
        "mappings":        mappings,
        "file_extra_cols": extra_cols,
    }


@router.post("/final-templates/normalize")
async def normalize_file(
    file:          UploadFile = File(...),
    template_name: str        = Form(...),
    mapping_json:  str        = Form(...),
):
    """Rearrange and rename file columns to match the template, apply header colors, return .xlsx."""
    tpl_path = _FINAL_TEMPLATES_DIR / f"{template_name}.json"
    if not tpl_path.exists():
        raise HTTPException(404, f"Template '{template_name}' not found.")

    tpl_data = json.loads(tpl_path.read_text(encoding="utf-8"))
    tpl_col_colors = {c["name"]: c["argb"] for c in tpl_data.get("columns", [])}

    try:
        user_mappings = json.loads(mapping_json)
        # Expected: [{template_col: str, file_col: str|None}, ...]
    except Exception:
        raise HTTPException(400, "Invalid mapping_json — must be a JSON array.")

    content = await file.read()
    fname = file.filename or "output.xlsx"

    try:
        df = _load_df(content, fname)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    # Build rename map and output column order
    rename_map: dict[str, str] = {}
    final_order: list[str] = []
    for m in user_mappings:
        tpl_col  = m.get("template_col")
        file_col = m.get("file_col")
        if not file_col or file_col not in df.columns:
            continue  # skip / missing
        if file_col != tpl_col:
            rename_map[file_col] = tpl_col
        final_order.append(tpl_col)

    if rename_map:
        df = df.rename(columns=rename_map)

    available = [c for c in final_order if c in df.columns]
    df = df[available]

    # Build Excel workbook with template header colors applied
    buf = io.BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.append(list(df.columns))

    for cell in ws[1]:
        col_name = cell.value
        argb = tpl_col_colors.get(col_name, "00000000")
        # Only apply solid fills for valid 8-char ARGB hex values that are not "no fill"
        if argb and re.match(r'^[0-9A-Fa-f]{8}$', argb) and argb != "00000000":
            try:
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=argb)
            except Exception:
                pass

    for row_tuple in df.itertuples(index=False, name=None):
        cleaned = [
            None if (v is None or (isinstance(v, float) and math.isnan(v))) else v
            for v in row_tuple
        ]
        ws.append(cleaned)

    wb.save(buf)
    buf.seek(0)

    stem = fname.rsplit(".", 1)[0] if "." in fname else fname
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{stem}_normalized.xlsx"'},
    )


# ── CRUD for final templates ──────────────────────────────────────────────────

@router.get("/final-templates")
def list_final_templates():
    """List all saved final output templates (summary, no columns array)."""
    result = []
    if _FINAL_TEMPLATES_DIR.exists():
        for f in sorted(_FINAL_TEMPLATES_DIR.glob("*.json")):
            try:
                data = json.loads(f.read_text(encoding="utf-8"))
                result.append({
                    "name":         f.stem,
                    "comment":      data.get("_comment", ""),
                    "column_count": len(data.get("columns", [])),
                    "created_at":   data.get("created_at", ""),
                })
            except Exception:
                pass
    return result


@router.get("/final-templates/{name}")
def get_final_template(name: str):
    """Get a single final template including full columns array."""
    path = _FINAL_TEMPLATES_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Template '{name}' not found.")
    data = json.loads(path.read_text(encoding="utf-8"))
    return {
        "name":       name,
        "comment":    data.get("_comment", ""),
        "created_at": data.get("created_at", ""),
        "columns":    data.get("columns", []),
    }


@router.post("/final-templates/{name}")
def save_final_template(name: str, payload: FinalTemplatePayload):
    """Save (create or overwrite) a final output template."""
    safe = "".join(c for c in name if c.isalnum() or c in "_-")
    if not safe:
        raise HTTPException(400, "Invalid template name.")
    if not payload.columns:
        raise HTTPException(400, "Template must have at least one column.")
    _FINAL_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    path = _FINAL_TEMPLATES_DIR / f"{safe}.json"
    data = {
        "_comment":   payload.comment or f"Final output template: {safe}",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "columns":    [{"name": c.name, "argb": c.argb} for c in payload.columns],
    }
    path.write_text(json.dumps(data, indent=4), encoding="utf-8")
    return {"name": safe, "status": "saved"}


@router.delete("/final-templates/{name}")
def delete_final_template(name: str):
    """Delete a final output template."""
    path = _FINAL_TEMPLATES_DIR / f"{name}.json"
    if not path.exists():
        raise HTTPException(404, f"Template '{name}' not found.")
    path.unlink()
    return {"status": "deleted"}
