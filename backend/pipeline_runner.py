"""
pipeline_runner.py — Executes all applicable QC steps autonomously.

Given a role_map (role → column) and a DataFrame, runs every step whose
required roles are present. Adds flag columns; never deletes source data.
"""
import sys
import os
import io
import traceback

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

# Ensure project root is importable
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)


def _has(role_map: dict, *roles) -> bool:
    """True if all given roles are present and non-empty in role_map."""
    for r in roles:
        val = role_map.get(r)
        if not val:
            return False
        if isinstance(val, list) and len(val) == 0:
            return False
    return True


def run_pipeline(df: pd.DataFrame, role_map: dict, progress_cb=None) -> tuple[pd.DataFrame, list, list]:
    """
    Run all applicable QC steps.
    Returns (df_result, step_results, columns_added).

    progress_cb: callable(message: str, pct: int)
    """
    results = []
    columns_added = []

    def _step(name: str, pct: int, fn):
        nonlocal df
        if progress_cb:
            progress_cb(f"Running: {name}", pct)
        try:
            before = set(df.columns)
            df, status = fn(df)
            after = set(df.columns)
            new_cols = [c for c in df.columns if c not in before]
            columns_added.extend(new_cols)
            results.append({"step": name, "status": status.get("status", "success"), "detail": status})
        except Exception as e:
            results.append({"step": name, "status": "error", "detail": {"error": str(e)}})

    rm = role_map
    ph = rm.get("phone_columns", [])

    total_steps = 27
    step_idx = [0]

    def next_pct():
        step_idx[0] += 1
        return int(step_idx[0] / total_steps * 80) + 10  # 10-90%

    # ── Email checks ────────────────────────────────────────────────────────
    if _has(rm, "email"):
        from functions_qc.email_handling.email_structure_validation import validate_email_structure
        _step("Email Structure", next_pct(), lambda d: validate_email_structure(d, rm["email"]))

        from functions_qc.email_handling.email_tld_check import check_email_tld
        _step("Email TLD", next_pct(), lambda d: check_email_tld(d, rm["email"]))

        from functions_qc.email_handling.email_disposable_check import check_disposable_email
        _step("Email Disposable", next_pct(), lambda d: check_disposable_email(d, rm["email"]))

        from functions_qc.email_handling.email_role_check import check_role_email
        _step("Email Role Account", next_pct(), lambda d: check_role_email(d, rm["email"]))

        from functions_qc.email_handling.email_reuse_check import check_email_reuse
        _step("Email Reuse", next_pct(), lambda d: check_email_reuse(d, rm["email"]))

    if _has(rm, "company", "email"):
        from functions_qc.email_handling.company_email_domain_match import company_email_domain_match
        _step("Company-Email Domain", next_pct(),
              lambda d: company_email_domain_match(d, rm["company"], rm["email"]))

    if _has(rm, "first_name", "last_name", "email"):
        from functions_qc.email_handling.name_email_fuzzy import name_email_fuzzy_match
        _step("Name-Email Fuzzy", next_pct(),
              lambda d: name_email_fuzzy_match(d, rm["first_name"], rm["last_name"], rm["email"],
                                               rm.get("middle_name")))

    # ── Name checks ─────────────────────────────────────────────────────────
    if _has(rm, "first_name") or _has(rm, "last_name"):
        from functions_qc.name_handling.null_name_check import check_null_names
        _step("Null Name Check", next_pct(),
              lambda d: check_null_names(d, rm.get("first_name"), rm.get("last_name")))

        from functions_qc.name_handling.dummy_names_check import check_dummy_names
        _step("Dummy Names", next_pct(),
              lambda d: check_dummy_names(d, rm.get("first_name"), rm.get("last_name")))

    if _has(rm, "first_name", "last_name"):
        from functions_qc.name_handling.non_alpha_name_handle import check_name_non_alphabetic_content
        _step("Non-Alpha Names", next_pct(),
              lambda d: check_name_non_alphabetic_content(d, rm["first_name"],
                                                           rm.get("middle_name", ""),
                                                           rm["last_name"]))

        from functions_qc.name_handling.dot_remove import remove_dot_from_names
        _step("Dot Remove", next_pct(),
              lambda d: remove_dot_from_names(d, rm["first_name"],
                                               rm.get("middle_name", ""),
                                               rm["last_name"]))

    if _has(rm, "first_name", "last_name", "company"):
        from functions_qc.name_handling.name_company_match import check_company_name_direct_match
        _step("Name-Company Match", next_pct(),
              lambda d: check_company_name_direct_match(d, rm["first_name"],
                                                         rm.get("middle_name", ""),
                                                         rm["last_name"], rm["company"]))

    if _has(rm, "full_name"):
        from functions_qc.name_handling.name_split import split_full_name
        _step("Name Split", next_pct(), lambda d: split_full_name(d, rm["full_name"]))

    # ── Phone checks ────────────────────────────────────────────────────────
    if ph:
        from functions_qc.phone_number_handling.standaize_phone_number import normalize_phone_excel
        for col in ph:
            col_snap = col
            _step(f"Standardize Phone ({col_snap})", next_pct(),
                  lambda d, c=col_snap: normalize_phone_excel(d, c))

        if _has(rm, "office_state"):
            from functions_qc.phone_number_handling.address_phone_postal import validate_phone_state
            for col in ph:
                col_snap = col
                _step(f"Phone-State Match ({col_snap})", next_pct(),
                      lambda d, c=col_snap: validate_phone_state(
                          d, f"{c}_country_code", f"{c}_standardized_number", rm["office_state"]))

        from functions_qc.phone_number_handling.reused_phone_check import check_reused_phone
        _step("Reused Phone", next_pct(), lambda d: check_reused_phone(d, ph))

    # ── Employee count ───────────────────────────────────────────────────────
    if _has(rm, "employee_count"):
        from functions_qc.employee_count_handler.employee_count_normalizer import normalize_employee_count
        _step("Employee Count Normalize", next_pct(),
              lambda d: normalize_employee_count(d, rm["employee_count"]))

    # ── LinkedIn ─────────────────────────────────────────────────────────────
    if _has(rm, "linkedin"):
        from functions_qc.linkedin_handler.linkedin_url_check import check_linkedin_url
        _step("LinkedIn URL Check", next_pct(), lambda d: check_linkedin_url(d, rm["linkedin"]))

        if _has(rm, "first_name", "last_name"):
            from functions_qc.linkedin_handler.name_linkedin_fuzzy_match import verify_linkedin_profile_match
            _step("LinkedIn Name Match", next_pct(),
                  lambda d: verify_linkedin_profile_match(d, rm["first_name"],
                                                           rm.get("middle_name", ""),
                                                           rm["last_name"], rm["linkedin"]))

    # ── Industry / Job / SIC ────────────────────────────────────────────────
    if _has(rm, "primary_industry"):
        from functions_qc.primary_industry.primary_industry_split_using_graterthen_greater_then import extract_primary_industry
        _step("Primary Industry Extract", next_pct(),
              lambda d: extract_primary_industry(d, rm["primary_industry"]))

    if _has(rm, "job_title"):
        from functions_qc.job_title_handler.job_title_categories import categorize_job_titles
        _step("Job Title Categorize", next_pct(),
              lambda d: categorize_job_titles(d, rm["job_title"]))

        from functions_qc.job_title_handler.job_title_non_alpha import check_job_title_non_alpha
        _step("Job Title Non-Alpha", next_pct(),
              lambda d: check_job_title_non_alpha(d, rm["job_title"]))

    if _has(rm, "sic_code"):
        from functions_qc.sic_code_naics_handler.sic_naics_handler import process_sic_naics
        _step("SIC → NAICS", next_pct(), lambda d: process_sic_naics(d, rm["sic_code"]))

    # ── Link text ───────────────────────────────────────────────────────────
    if _has(rm, "company", "link_text", "description") and (
        _has(rm, "first_name", "last_name") or _has(rm, "full_name")
    ):
        from functions_qc.link_text_handler.link_text_match import match_link_text_fields
        _step("Link Text Match", next_pct(),
              lambda d: match_link_text_fields(d, rm["company"], rm["link_text"], rm["description"],
                                               full_name_col=rm.get("full_name"),
                                               first_name_col=rm.get("first_name"),
                                               last_name_col=rm.get("last_name"),
                                               middle_name_col=rm.get("middle_name")))

    # ── Facebook ─────────────────────────────────────────────────────────────
    if _has(rm, "first_name", "last_name") and (
        _has(rm, "facebook") or _has(rm, "facebook_link_text") or _has(rm, "facebook_description")
    ):
        from functions_qc.facebook_handler.facebook_match import verify_facebook_profile_match
        extra_fb = [c for c in [rm.get("facebook_link_text"), rm.get("facebook_description")] if c]
        _step("Facebook Name Match", next_pct(),
              lambda d: verify_facebook_profile_match(d, rm["first_name"], rm["last_name"],
                                                       middle_name_col=rm.get("middle_name"),
                                                       facebook_col=rm.get("facebook"),
                                                       extra_cols=extra_fb or None))

    # ── Revenue / City-State-Postal ─────────────────────────────────────────
    if _has(rm, "company_revenue"):
        from functions_qc.company_revenue_handler.company_revenue_check import check_company_revenue
        _step("Company Revenue Check", next_pct(),
              lambda d: check_company_revenue(d, rm["company_revenue"]))

    if _has(rm, "city") and _has(rm, "postal_code"):
        from functions_qc.city_state_postal_handler.city_state_postal_match import check_city_state_postal
        state_col = rm.get("state") or rm.get("office_state") or ""
        _step("City-State-Postal Match", next_pct(),
              lambda d: check_city_state_postal(d, rm["postal_code"], state_col, rm["city"]))

    if progress_cb:
        progress_cb("Generating Excel output", 92)

    return df, results, columns_added


def build_excel_bytes(df: pd.DataFrame, original_columns: list) -> bytes:
    """Write DataFrame to Excel bytes with QC columns highlighted purple."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="QC_Output")
        ws = writer.sheets["QC_Output"]

        purple_fill = PatternFill(start_color="D8B4FE", end_color="D8B4FE", fill_type="solid")
        orig_set = set(original_columns)

        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name not in orig_set:
                ws.cell(row=1, column=col_idx).fill = purple_fill

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return buf.getvalue()
